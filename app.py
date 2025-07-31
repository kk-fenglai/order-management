import threading
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
import os
import qrcode
from io import BytesIO
import base64
import logging
from config import config
from config_local import *
from models import db, Package
from utils import send_shenzhen_arrival_email, send_cafe_arrival_email, generate_pickup_codes_qr, validate_email_address
from functools import wraps

def send_email_async(package, email_type, mail, app):
    """异步发送邮件"""
    with app.app_context():
        try:
            if email_type == 'shenzhen':
                success = send_shenzhen_arrival_email(package, mail)
            elif email_type == 'cafe':
                success = send_cafe_arrival_email(package, mail)
            else:
                return False
            
            if success:
                logging.info(f'异步邮件发送成功: {package.customer_email} ({email_type})')
            else:
                logging.error(f'异步邮件发送失败: {package.customer_email} ({email_type})')
            
            return success
        except Exception as e:
            logging.error(f'异步邮件发送异常: {str(e)}')
            return False

def create_app(config_name='default'):
    """应用工厂函数"""
    app = Flask(__name__)
    app.config.from_object(config[config_name])
    
    # 初始化扩展
    db.init_app(app)
    mail = Mail(app)
    
    # 注册模板过滤器
    @app.template_filter('format_datetime')
    def format_datetime_filter(dt):
        if dt is None:
            return ''
        # 转换为巴黎时间
        try:
            import pytz
            utc_tz = pytz.UTC
            paris_tz = pytz.timezone('Europe/Paris')
            if dt.tzinfo is None:
                dt = utc_tz.localize(dt)
            paris_time = dt.astimezone(paris_tz)
            return paris_time.strftime('%Y-%m-%d %H:%M:%S')
        except:
            # 如果转换失败，返回原始时间
            return dt.strftime('%Y-%m-%d %H:%M:%S')
    
    @app.template_filter('get_status_info')
    def get_status_info_filter(status):
        status_info = {
            'shenzhen_arrived': {'display': '已到深圳', 'icon': 'fas fa-truck', 'class': 'primary'},
            'cafe_arrived': {'display': '已到咖啡馆', 'icon': 'fas fa-store', 'class': 'warning'},
            'picked_up': {'display': '已取件', 'icon': 'fas fa-check-circle', 'class': 'success'}
        }
        return status_info.get(status, {'display': '未知状态', 'icon': 'fas fa-question', 'class': 'secondary'})
    
    # 错误处理
    @app.errorhandler(404)
    def not_found_error(error):
        return render_template('errors/404.html'), 404
    
    @app.errorhandler(500)
    def internal_error(error):
        db.session.rollback()
        return render_template('errors/500.html'), 500
    
    # 路由
    @app.route('/', methods=['GET', 'POST'])
    def index():
        """主页 - 显示所有包裹，支持Excel导入"""
        import_result = None
        shenzhen_email_result = None
        cafe_email_result = None
        
        # 从session获取邮件发送结果
        if 'shenzhen_email_result' in session:
            shenzhen_email_result = session.pop('shenzhen_email_result')
        if 'cafe_email_result' in session:
            cafe_email_result = session.pop('cafe_email_result')
        
        # 处理Excel导入
        if request.method == 'POST':
            file = request.files.get('excel_file')
            if not file:
                import_result = {'success': False, 'error': '请选择要上传的文件'}
            elif not file.filename:
                import_result = {'success': False, 'error': '文件名不能为空'}
            elif not file.filename.endswith('.xlsx'):
                import_result = {'success': False, 'error': '请上传xlsx格式的Excel文件'}
            else:
                try:
                    # 检查文件是否为空
                    file.seek(0, 2)  # 移动到文件末尾
                    file_size = file.tell()
                    file.seek(0)  # 重置到文件开头
                    
                    if file_size == 0:
                        import_result = {'success': False, 'error': '上传的文件为空'}
                    elif file_size > app.config.get('MAX_CONTENT_LENGTH', 16 * 1024 * 1024):
                        import_result = {'success': False, 'error': '文件大小超过限制（最大16MB）'}
                    else:
                        # 使用openpyxl读取Excel文件
                        from openpyxl import load_workbook
                        wb = load_workbook(file, data_only=True)
                        ws = wb.active
                        
                        # 获取表头
                        headers = []
                        for cell in ws[1]:
                            headers.append(str(cell.value).strip() if cell.value else '')
                        
                        required_cols = ['客户名称', '快递单号', '客户邮箱', '备注']
                        missing_cols = [col for col in required_cols if col not in headers]
                        
                        if missing_cols:
                            import_result = {'success': False, 'error': f'Excel缺少必要列: {", ".join(missing_cols)}，请使用模板'}
                        else:
                            rows = []
                            ids = []
                            skipped_rows = []
                            
                            # 获取列索引
                            customer_name_idx = headers.index('客户名称')
                            tracking_number_idx = headers.index('快递单号')
                            email_idx = headers.index('客户邮箱')
                            notes_idx = headers.index('备注')
                            
                            # 处理数据行
                            for row_num in range(2, ws.max_row + 1):
                                customer_name = str(ws.cell(row=row_num, column=customer_name_idx + 1).value or '').strip()
                                tracking_number = str(ws.cell(row=row_num, column=tracking_number_idx + 1).value or '').strip()
                                email = str(ws.cell(row=row_num, column=email_idx + 1).value or '').strip()
                                notes = str(ws.cell(row=row_num, column=notes_idx + 1).value or '').strip()
                                
                                if not customer_name or not tracking_number or not email:
                                    continue
                                
                                # 检查快递单号是否已存在
                                existing_package = Package.query.filter_by(shenzhen_tracking_number=tracking_number).first()
                                if existing_package:
                                    skipped_rows.append({
                                        'customer_name': customer_name,
                                        'tracking_number': tracking_number,
                                        'email': email,
                                        'notes': notes,
                                        'reason': '快递单号已存在'
                                    })
                                    continue
                                
                                # 创建包裹
                                package = Package(
                                    customer_name=customer_name,
                                    shenzhen_tracking_number=tracking_number,
                                    customer_email=email,
                                    notes=notes,
                                    pickup_code=Package.generate_pickup_code()
                                )
                                db.session.add(package)
                                db.session.flush()  # 获取ID
                                ids.append(package.id)
                                rows.append({
                                    'customer_name': customer_name,
                                    'tracking_number': tracking_number,
                                    'email': email,
                                    'notes': notes
                                })
                            db.session.commit()
                            import_result = {
                                'success': True, 
                                'count': len(rows), 
                                'rows': rows, 
                                'ids': ids,
                                'skipped_count': len(skipped_rows),
                                'skipped_rows': skipped_rows
                            }
                except Exception as e:
                    db.session.rollback()
                    error_msg = str(e)
                    if "File is not a zip file" in error_msg:
                        import_result = {'success': False, 'error': '文件格式错误：请确保上传的是有效的Excel文件(.xlsx)，不是其他格式的文件'}
                    elif "No sheet names" in error_msg:
                        import_result = {'success': False, 'error': 'Excel文件没有工作表，请检查文件内容'}
                    else:
                        import_result = {'success': False, 'error': f'导入失败: {error_msg}'}
        
        # 获取包裹列表
        page = request.args.get('page', 1, type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '')
        
        # 构建查询
        query = Package.query
        
        if status_filter:
            query = query.filter_by(status=status_filter)
        
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    Package.customer_name.like(search_term),
                    Package.customer_email.like(search_term),
                    Package.shenzhen_tracking_number.like(search_term),
                    Package.pickup_code.like(search_term)
                )
            )
        
        # 分页
        packages = query.order_by(Package.cafe_arrival_date.desc().nullslast()).paginate(
            page=page, per_page=app.config['ITEMS_PER_PAGE'], error_out=False
        )
        
        return render_template('index.html', packages=packages, 
                             status_filter=status_filter, search=search,
                             import_result=import_result, 
                             shenzhen_email_result=shenzhen_email_result,
                             cafe_email_result=cafe_email_result)
    
    @app.route('/new_package', methods=['GET', 'POST'])
    def new_package():
        """创建新包裹记录"""
        if request.method == 'POST':
            # 获取表单数据
            customer_name = request.form.get('customer_name', '').strip()
            customer_email = request.form.get('customer_email', '').strip()
            shenzhen_tracking_number = request.form.get('shenzhen_tracking_number', '').strip()
            notes = request.form.get('notes', '').strip()
            
            # 输入验证
            errors = []
            if not customer_name:
                errors.append('客户姓名不能为空')
            if not customer_email:
                errors.append('客户邮箱不能为空')
            elif not validate_email_address(customer_email):
                errors.append('邮箱格式不正确')
            if not shenzhen_tracking_number:
                errors.append('深圳快递单号不能为空')
            
            if errors:
                for error in errors:
                    flash(error, 'error')
                return render_template('new_package.html', 
                                     customer_name=customer_name,
                                     customer_email=customer_email,
                                     shenzhen_tracking_number=shenzhen_tracking_number,
                                     notes=notes)
            
            # 检查深圳快递单号是否已存在
            existing_package = Package.query.filter_by(
                shenzhen_tracking_number=shenzhen_tracking_number
            ).first()
            if existing_package:
                flash('深圳快递单号已存在，请检查后重新输入', 'error')
                return render_template('new_package.html',
                                     customer_name=customer_name,
                                     customer_email=customer_email,
                                     shenzhen_tracking_number=shenzhen_tracking_number,
                                     notes=notes)
            
            try:
                # 创建新包裹记录
                package = Package(
                    customer_name=customer_name,
                    customer_email=customer_email,
                    shenzhen_tracking_number=shenzhen_tracking_number,
                    pickup_code=Package.generate_pickup_code(),
                    notes=notes
                )
                
                db.session.add(package)
                db.session.commit()
                
                # 发送深圳到达通知邮件
                if send_shenzhen_arrival_email(package, mail):
                    flash(f'包裹记录创建成功！深圳快递单号: {shenzhen_tracking_number}，取件码: {package.pickup_code}，深圳到达通知邮件已发送给客户。', 'success')
                else:
                    flash(f'包裹记录创建成功！深圳快递单号: {shenzhen_tracking_number}，取件码: {package.pickup_code}，但邮件发送失败。请检查邮件配置。', 'warning')
                
                return redirect(url_for('index'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'创建包裹记录失败: {str(e)}', 'error')
                return render_template('new_package.html',
                                     customer_name=customer_name,
                                     customer_email=customer_email,
                                     shenzhen_tracking_number=shenzhen_tracking_number,
                                     notes=notes)
        
        return render_template('new_package.html')
    
    @app.route('/package/<int:package_id>')
    def package_detail(package_id):
        """包裹详情页面"""
        package = Package.query.get_or_404(package_id)
        return render_template('package_detail.html', package=package)
    
    @app.route('/package/<int:package_id>/resend_shenzhen_email')
    def resend_shenzhen_email(package_id):
        """重新发送深圳到达通知邮件"""
        package = Package.query.get_or_404(package_id)
        
        if send_shenzhen_arrival_email(package, mail):
            flash('深圳到达通知邮件已重新发送！', 'success')
        else:
            flash('邮件发送失败。请检查邮件配置。', 'error')
        
        return redirect(url_for('package_detail', package_id=package_id))
    
    @app.route('/package/<int:package_id>/resend_cafe_email')
    def resend_cafe_email(package_id):
        """重新发送咖啡馆到达通知邮件"""
        package = Package.query.get_or_404(package_id)
        
        if send_cafe_arrival_email(package, mail):
            flash('咖啡馆到达通知邮件已重新发送！', 'success')
        else:
            flash('邮件发送失败。请检查邮件配置。', 'error')
        
        return redirect(url_for('package_detail', package_id=package_id))
    
    @app.route('/package/<int:package_id>/update_status', methods=['POST'])
    def update_status(package_id):
        """更新包裹状态"""
        package = Package.query.get_or_404(package_id)
        new_status = request.form.get('status')
        
        if new_status not in ['shenzhen_arrived', 'cafe_arrived', 'picked_up']:
            flash('无效的状态值', 'error')
            return redirect(url_for('package_detail', package_id=package_id))
        
        try:
            package.status = new_status
            
            # 如果标记为已取件，设置取件时间
            if new_status == 'picked_up' and not package.pickup_date:
                package.pickup_date = datetime.utcnow()
            
            # 如果标记为到达咖啡馆，设置到达时间
            if new_status == 'cafe_arrived' and not package.cafe_arrival_date:
                package.cafe_arrival_date = datetime.utcnow()
            
            package.updated_at = datetime.utcnow()
            db.session.commit()
            
            status_info = {
                'shenzhen_arrived': '已到深圳',
                'cafe_arrived': '已到咖啡馆', 
                'picked_up': '已取件'
            }
            flash(f'包裹状态已更新为: {status_info.get(new_status, "未知状态")}', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'更新状态失败: {str(e)}', 'error')
        
        return redirect(url_for('package_detail', package_id=package_id))
    
    @app.route('/package/<int:package_id>/cafe_arrival')
    def cafe_arrival(package_id):
        """标记包裹到达咖啡馆并发送取件码邮件"""
        package = Package.query.get_or_404(package_id)
        
        if package.status != 'shenzhen_arrived':
            flash('包裹必须先到达深圳仓库', 'error')
            return redirect(url_for('package_detail', package_id=package_id))
        
        try:
            # 更新状态和到达时间
            package.status = 'cafe_arrived'
            package.cafe_arrival_date = datetime.utcnow()
            package.updated_at = datetime.utcnow()
            db.session.commit()
            
            # 发送咖啡馆到达通知邮件
            if send_cafe_arrival_email(package, mail):
                flash(f'包裹已标记为到达咖啡馆！取件码: {package.pickup_code}，取件通知邮件已发送给客户。', 'success')
            else:
                flash(f'包裹已标记为到达咖啡馆！取件码: {package.pickup_code}，但邮件发送失败。请检查邮件配置。', 'warning')
            
        except Exception as e:
            db.session.rollback()
            flash(f'更新状态失败: {str(e)}', 'error')
        
        return redirect(url_for('package_detail', package_id=package_id))
    
    @app.route('/api/packages')
    def api_packages():
        """API接口 - 获取所有包裹"""
        packages = Package.query.all()
        return jsonify([package.to_dict() for package in packages])
    
    @app.route('/api/pickup_codes')
    def api_pickup_codes():
        """API接口 - 获取待取件包裹的取件码"""
        packages = Package.query.filter_by(status='cafe_arrived').order_by(Package.cafe_arrival_date.desc()).all()
        pickup_data = []
        
        for package in packages:
            pickup_data.append({
                'id': package.id,
                'pickup_code': package.pickup_code,
                'customer_name': package.customer_name,
                'shenzhen_tracking_number': package.shenzhen_tracking_number,
                'cafe_arrival_date': package.cafe_arrival_date_paris.strftime('%m-%d %H:%M') if package.cafe_arrival_date_paris else None,
                'latest_pickup_time': package.latest_pickup_time_paris.strftime('%m-%d %H:%M') if package.latest_pickup_time_paris else None,
                'is_overdue': package.is_overdue
            })
        
        return jsonify({
            'success': True,
            'data': pickup_data,
            'count': len(pickup_data),
            'timestamp': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        })
    
    @app.route('/qr_codes')
    def qr_codes():
        """二维码页面 - 显示所有取件码的二维码"""
        # 获取所有有取件码的包裹
        packages = Package.query.filter(Package.pickup_code.isnot(None)).order_by(Package.cafe_arrival_date.desc().nullslast()).all()
        
        if not packages:
            flash('目前没有包裹', 'info')
            return redirect(url_for('index'))
        
        qr_image, qr_text = generate_pickup_codes_qr(packages)
        
        # 提取URL用于显示
        qr_url = qr_text.split('URL: ')[1].split('\n')[0] if 'URL: ' in qr_text else qr_text
        
        return render_template('qr_codes.html', 
                             qr_image=qr_image, 
                             qr_text=qr_text,
                             qr_url=qr_url,
                             packages=packages)
    
    @app.route('/mobile_pickup')
    def mobile_pickup():
        """移动端取件码页面 - 方便店家在手机上查看"""
        # 获取所有有取件码的包裹
        packages = Package.query.filter(Package.pickup_code.isnot(None)).order_by(Package.cafe_arrival_date.desc().nullslast()).all()
        return render_template('mobile_pickup.html', packages=packages)
    
    @app.route('/pickup_cards')
    def pickup_cards():
        """取件码卡片页面 - 方便打印单个取件码"""
        # 获取已到达咖啡馆的包裹
        packages = Package.query.filter_by(status='cafe_arrived').order_by(Package.cafe_arrival_date.desc()).all()
        return render_template('pickup_cards.html', packages=packages)
    
    @app.route('/stats')
    def stats():
        """统计信息页面"""
        total_packages = Package.query.count()
        shenzhen_arrived = Package.query.filter_by(status='shenzhen_arrived').count()
        cafe_arrived = Package.query.filter_by(status='cafe_arrived').count()
        picked_up = Package.query.filter_by(status='picked_up').count()
        
        # 今日数据
        today = datetime.utcnow().date()
        today_packages = Package.query.filter(
            db.func.date(Package.cafe_arrival_date) == today
        ).count()
        
        stats_data = {
            'total': total_packages,
            'shenzhen_arrived': shenzhen_arrived,
            'cafe_arrived': cafe_arrived,
            'picked_up': picked_up,
            'today': today_packages
        }
        
        return render_template('stats.html', stats=stats_data)
    
    @app.route('/import_template.xlsx')
    def download_import_template():
        return send_from_directory('.', 'import_template.xlsx', as_attachment=True)

    @app.route('/import_excel', methods=['GET', 'POST'])
    def import_excel():
        import_result = None
        if request.method == 'POST':
            file = request.files.get('excel_file')
            if not file:
                import_result = {'success': False, 'error': '请选择要上传的文件'}
                return render_template('import_excel.html', import_result=import_result)
            
            if not file.filename:
                import_result = {'success': False, 'error': '文件名不能为空'}
                return render_template('import_excel.html', import_result=import_result)
            
            if not file.filename.endswith('.xlsx'):
                import_result = {'success': False, 'error': '请上传xlsx格式的Excel文件'}
                return render_template('import_excel.html', import_result=import_result)
            
            try:
                # 检查文件是否为空
                file.seek(0, 2)  # 移动到文件末尾
                file_size = file.tell()
                file.seek(0)  # 重置到文件开头
                
                if file_size == 0:
                    import_result = {'success': False, 'error': '上传的文件为空'}
                    return render_template('import_excel.html', import_result=import_result)
                
                # 使用openpyxl读取Excel文件
                from openpyxl import load_workbook
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                
                # 获取表头
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).strip() if cell.value else '')
                
                required_cols = ['客户名称', '快递单号', '客户邮箱', '备注']
                missing_cols = [col for col in required_cols if col not in headers]
                
                if missing_cols:
                    import_result = {'success': False, 'error': f'Excel缺少必要列: {", ".join(missing_cols)}，请使用模板'}
                    return render_template('import_excel.html', import_result=import_result)
                
                rows = []
                ids = []
                skipped_rows = []
                
                # 获取列索引
                customer_name_idx = headers.index('客户名称')
                tracking_number_idx = headers.index('快递单号')
                email_idx = headers.index('客户邮箱')
                notes_idx = headers.index('备注')
                
                # 处理数据行
                for row_num in range(2, ws.max_row + 1):
                    customer_name = str(ws.cell(row=row_num, column=customer_name_idx + 1).value or '').strip()
                    tracking_number = str(ws.cell(row=row_num, column=tracking_number_idx + 1).value or '').strip()
                    email = str(ws.cell(row=row_num, column=email_idx + 1).value or '').strip()
                    notes = str(ws.cell(row=row_num, column=notes_idx + 1).value or '').strip()
                    
                    if not customer_name or not tracking_number or not email:
                        continue
                    
                    # 检查快递单号是否已存在
                    existing_package = Package.query.filter_by(shenzhen_tracking_number=tracking_number).first()
                    if existing_package:
                        skipped_rows.append({
                            'customer_name': customer_name,
                            'tracking_number': tracking_number,
                            'email': email,
                            'notes': notes,
                            'reason': '快递单号已存在'
                        })
                        continue
                    
                    # 创建包裹
                    package = Package(
                        customer_name=customer_name,
                        shenzhen_tracking_number=tracking_number,
                        customer_email=email,
                        notes=notes,
                        pickup_code=Package.generate_pickup_code()  # 生成取件码
                    )
                    db.session.add(package)
                    db.session.flush()  # 获取ID
                    ids.append(package.id)
                    rows.append({
                        'customer_name': customer_name,
                        'tracking_number': tracking_number,
                        'email': email,
                        'notes': notes
                    })
                db.session.commit()
                import_result = {
                    'success': True, 
                    'count': len(rows), 
                    'rows': rows, 
                    'ids': ids,
                    'skipped_count': len(skipped_rows),
                    'skipped_rows': skipped_rows
                }
            except Exception as e:
                db.session.rollback()
                error_msg = str(e)
                if "File is not a zip file" in error_msg:
                    import_result = {'success': False, 'error': '文件格式错误：请确保上传的是有效的Excel文件(.xlsx)，不是其他格式的文件'}
                elif "No sheet names" in error_msg:
                    import_result = {'success': False, 'error': 'Excel文件没有工作表，请检查文件内容'}
                else:
                    import_result = {'success': False, 'error': f'导入失败: {error_msg}'}
        return render_template('import_excel.html', import_result=import_result)

    @app.route('/import_excel_send', methods=['POST'])
    def import_excel_send():
        ids = request.form.get('import_ids', '')
        id_list = [int(i) for i in ids.split(',') if i.isdigit()]
        packages = Package.query.filter(Package.id.in_(id_list)).all()
        mail = Mail(app)
        success, fail = 0, 0
        for package in packages:
            if send_shenzhen_arrival_email(package, mail):
                success += 1
            else:
                fail += 1
        flash(f'群发完成，成功 {success} 条，失败 {fail} 条', 'success' if fail == 0 else 'warning')
        return redirect(url_for('import_excel'))

    @app.route('/send_shenzhen_emails', methods=['POST'])
    def send_shenzhen_emails():
        """发送所有待发送的深圳到达邮件"""
        try:
            mail = Mail(app)
            count = 0
            
            # 查找需要发送深圳到达邮件的包裹
            shenzhen_packages = Package.query.filter_by(
                status='shenzhen_arrived',
                shenzhen_email_sent=False
            ).all()
            
            # 异步发送所有邮件
            def send_bulk_shenzhen_emails():
                nonlocal count
                for package in shenzhen_packages:
                    if send_shenzhen_arrival_email(package, mail):
                        count += 1
                logging.info(f'批量深圳邮件发送完成，成功: {count} 条')
            
            thread = threading.Thread(target=send_bulk_shenzhen_emails)
            thread.daemon = True
            thread.start()
            
            # 立即返回结果
            session['shenzhen_email_result'] = {
                'success': True,
                'count': len(shenzhen_packages),
                'message': f'正在后台发送 {len(shenzhen_packages)} 条深圳邮件'
            }
            
            return redirect(url_for('index'))
            
        except Exception as e:
            session['shenzhen_email_result'] = {
                'success': False,
                'error': f'发送深圳邮件时发生错误: {str(e)}'
            }
            return redirect(url_for('index'))

    @app.route('/send_cafe_emails', methods=['POST'])
    def send_cafe_emails():
        """发送所有待发送的咖啡馆到达邮件"""
        try:
            mail = Mail(app)
            count = 0
            
            # 查找需要发送咖啡馆到达邮件的包裹
            cafe_packages = Package.query.filter_by(
                status='cafe_arrived',
                cafe_email_sent=False
            ).all()
            
            # 异步发送所有邮件
            def send_bulk_cafe_emails():
                nonlocal count
                for package in cafe_packages:
                    if send_cafe_arrival_email(package, mail):
                        count += 1
                logging.info(f'批量咖啡馆邮件发送完成，成功: {count} 条')
            
            thread = threading.Thread(target=send_bulk_cafe_emails)
            thread.daemon = True
            thread.start()
            
            # 立即返回结果
            session['cafe_email_result'] = {
                'success': True,
                'count': len(cafe_packages),
                'message': f'正在后台发送 {len(cafe_packages)} 条咖啡馆邮件'
            }
            
            return redirect(url_for('index'))
            
        except Exception as e:
            session['cafe_email_result'] = {
                'success': False,
                'error': f'发送咖啡馆邮件时发生错误: {str(e)}'
            }
            return redirect(url_for('index'))
    
    @app.route('/package/<int:package_id>/delete', methods=['POST'])
    def delete_package(package_id):
        """删除包裹"""
        try:
            package = Package.query.get_or_404(package_id)
            customer_name = package.customer_name
            tracking_number = package.shenzhen_tracking_number
            
            db.session.delete(package)
            db.session.commit()
            
            flash(f'包裹已删除：客户 {customer_name}，快递单号 {tracking_number}', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'删除包裹失败: {str(e)}', 'error')
            return redirect(url_for('package_detail', package_id=package_id))
    
    @app.route('/delete_all_packages', methods=['POST'])
    def delete_all_packages():
        """删除所有包裹（危险操作）"""
        try:
            # 获取删除条件
            status_filter = request.form.get('status_filter', '')
            confirm_text = request.form.get('confirm_text', '')
            
            # 验证确认文本
            if confirm_text != 'DELETE ALL':
                flash('确认文本不正确，请输入 "DELETE ALL" 来确认删除', 'error')
                return redirect(url_for('index'))
            
            # 构建查询
            query = Package.query
            if status_filter:
                query = query.filter_by(status=status_filter)
            
            # 获取要删除的包裹数量
            packages_to_delete = query.all()
            count = len(packages_to_delete)
            
            if count == 0:
                flash('没有找到要删除的包裹', 'info')
                return redirect(url_for('index'))
            
            # 删除包裹
            for package in packages_to_delete:
                db.session.delete(package)
            
            db.session.commit()
            
            status_text = f'状态为 {status_filter} 的' if status_filter else '所有'
            flash(f'已删除 {count} 个{status_text}包裹', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'删除包裹失败: {str(e)}', 'error')
            return redirect(url_for('index'))
    
    @app.route('/package/<int:package_id>/quick_update', methods=['POST'])
    def quick_update_status(package_id):
        """快速更新包裹状态（用于首页）"""
        try:
            package = Package.query.get_or_404(package_id)
            new_status = request.form.get('status')
            
            if new_status not in ['cafe_arrived', 'picked_up']:
                flash('无效的状态值', 'error')
                return redirect(url_for('index'))
            
            # 更新状态
            old_status = package.status
            package.status = new_status
            
            # 如果是标记为到达咖啡馆
            if new_status == 'cafe_arrived' and old_status != 'cafe_arrived':
                package.cafe_arrival_date = datetime.utcnow()
                # 同步发送咖啡馆到达邮件
                mail = Mail(app)
                success = send_cafe_arrival_email(package, mail)
                if success:
                    flash(f'✅ 包裹已标记为到达咖啡馆！取件码: {package.pickup_code}，取件通知邮件发送成功。', 'success')
                else:
                    flash(f'⚠️ 包裹已标记为到达咖啡馆！取件码: {package.pickup_code}，但邮件发送失败。', 'warning')
            
            # 如果是标记为已取件
            elif new_status == 'picked_up' and old_status != 'picked_up':
                package.pickup_date = datetime.utcnow()
                flash(f'包裹已标记为已取件！客户: {package.customer_name}', 'success')
            
            db.session.commit()
            
        except Exception as e:
            db.session.rollback()
            flash(f'更新状态失败: {str(e)}', 'error')
        
        return redirect(url_for('index'))
    
    @app.route('/package/<int:package_id>/send_shenzhen_email', methods=['POST'])
    def send_single_shenzhen_email(package_id):
        """发送单个包裹的深圳到达邮件（用于首页）"""
        try:
            package = Package.query.get_or_404(package_id)
            
            if package.status != 'shenzhen_arrived':
                flash('只有已到深圳的包裹才能发送深圳到达邮件', 'warning')
                return redirect(url_for('index'))
            
            if package.shenzhen_email_sent:
                flash('深圳到达邮件已经发送过了', 'info')
                return redirect(url_for('index'))
            
            # 同步发送邮件（立即更新状态）
            mail = Mail(app)
            success = send_shenzhen_arrival_email(package, mail)
            
            if success:
                flash(f'✅ 深圳到达邮件发送成功！客户: {package.customer_name}', 'success')
            else:
                flash(f'❌ 深圳到达邮件发送失败！客户: {package.customer_name}', 'error')
            
        except Exception as e:
            flash(f'发送邮件时发生错误: {str(e)}', 'error')
        
        return redirect(url_for('index'))
    
    @app.route('/package/<int:package_id>/send_cafe_email', methods=['POST'])
    def send_single_cafe_email(package_id):
        """发送单个包裹的咖啡馆到达邮件"""
        package = Package.query.get_or_404(package_id)
        
        if package.status == 'shenzhen_arrived':
            # 更新状态为咖啡馆到达
            package.status = 'cafe_arrived'
            package.cafe_arrival_date = datetime.utcnow() # Changed from cafe_arrival_time to cafe_arrival_date
            db.session.commit()
            
            # 发送邮件
            success = send_cafe_arrival_email(package, mail)
            
            if success:
                flash('咖啡馆到达邮件发送成功！', 'success')
            else:
                flash('咖啡馆到达邮件发送失败，请检查邮箱配置。', 'error')
        else:
            flash('只有已到深圳的包裹才能发送咖啡馆到达邮件。', 'error')
        
        return redirect(url_for('package_detail', package_id=package_id))

    @app.route('/excel_converter', methods=['GET', 'POST'])
    def excel_converter():
        """Excel格式转换功能"""
        if request.method == 'POST':
            if 'file' not in request.files:
                flash('请选择要转换的Excel文件', 'error')
                return render_template('excel_converter.html')
            
            file = request.files['file']
            if file.filename == '':
                flash('请选择要转换的Excel文件', 'error')
                return render_template('excel_converter.html')
            
            if file and file.filename.endswith('.xlsx'):
                try:
                    import pandas as pd
                    from io import BytesIO
                    import base64
                    
                    # 读取原始Excel文件
                    df_original = pd.read_excel(file)
                    
                    # 创建转换后的数据框
                    converted_data = []
                    
                    # 遍历原始数据的每一行
                    for index, row in df_original.iterrows():
                        # 处理货物1-10列，提取快递单号
                        tracking_numbers = []
                        for i in range(1, 11):
                            col_name = f'货物{i}'
                            if col_name in row and pd.notna(row[col_name]) and str(row[col_name]).strip():
                                value = str(row[col_name])
                                # 提取快递单号（处理多种格式）
                                if '快递单号:' in value:
                                    # 处理标准格式: 快递单号:快递单号:实际单号
                                    if value.count('快递单号:') == 2:
                                        tracking_number = value.replace('快递单号:快递单号:', '').strip()
                                        # 移除可能的后缀（如",中文"）
                                        if ',' in tracking_number:
                                            tracking_number = tracking_number.split(',')[0].strip()
                                        if tracking_number and tracking_number != '快递单号:':
                                            tracking_numbers.append(tracking_number)
                                    # 处理重复前缀格式: 快递单号:快递单号:快递单号:...
                                    elif value.count('快递单号:') > 2:
                                        # 这种情况表示没有找到有效的快递单号，跳过
                                        continue
                        
                        # 为每个快递单号创建一行数据
                        for tracking_number in tracking_numbers:
                            converted_row = {
                                '客户姓名': row.get('姓名+拼音', '') if pd.notna(row.get('姓名+拼音', '')) else '',
                                '快递单号': tracking_number,
                                '中文品名 (必填)': '待填写',
                                '英文品名 (必填)': 'To be filled',
                                '材质 (中英文填写) (必填)': '待填写',
                                '用途 (中英文填写) (必填)': '待填写',
                                '国外海关编码 (必填)': '待填写',
                                '产品类型/属性 (必填)': '待填写',
                                '单位 套/包/组/个 (必填)': '个',
                                '单箱产品数量 单位:套/组/个 (必填)': 1,
                                '箱数/件数 CTN (必填为1)': 1,
                                '申报单价币种 请根据渠道选择 (必填)': 'EUR',
                                '单个产品申报单价 (仅保留小数点后两位且四舍五入) (必填)': 0.00,
                                '单个产品净重KG (选填) (仅保留小数点后两位且四舍五入)': 0.00,
                                '产品高清图片 (必须缩在方格内) (必填)': '',
                                '品牌 (如实申报)': '',
                                '品牌类型 (如实申报)': '',
                                '型号 (如实申报)': '',
                                '销售链接 (填写链接的前后不能有空格) (必填)': '',
                                '备注 (产品的其他特殊说明)': f'原始数据: 客户{row.get("姓名+拼音", "")}, 手机号{row.get("国手机号", "")}, 邮箱{row.get("邮箱", "")}',
                                '总产品数量': 1,
                                '总申报金额': 0.00
                            }
                            converted_data.append(converted_row)
                    
                    if not converted_data:
                        flash('未找到有效的快递单号数据', 'error')
                        return render_template('excel_converter.html')
                    
                    # 创建转换后的DataFrame
                    df_converted = pd.DataFrame(converted_data)
                    
                    # 生成转换后的Excel文件
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_converted.to_excel(writer, sheet_name='转换结果', index=False)
                    
                    output.seek(0)
                    
                    # 将文件内容编码为base64
                    file_content = base64.b64encode(output.getvalue()).decode()
                    
                    flash(f'转换成功！共转换 {len(converted_data)} 条记录', 'success')
                    
                    return render_template('excel_converter.html', 
                                         file_content=file_content, 
                                         filename=f'converted_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                                         converted_data=converted_data[:10])  # 只显示前10条预览
                    
                except Exception as e:
                    flash(f'转换失败: {str(e)}', 'error')
                    return render_template('excel_converter.html')
            else:
                flash('请上传.xlsx格式的Excel文件', 'error')
                return render_template('excel_converter.html')
        
        return render_template('excel_converter.html')

    @app.route('/download_converted_excel')
    def download_converted_excel():
        """下载转换后的Excel文件"""
        file_content = request.args.get('content')
        filename = request.args.get('filename', 'converted.xlsx')
        
        if file_content:
            try:
                import base64
                file_data = base64.b64decode(file_content)
                return send_file(
                    BytesIO(file_data),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            except Exception as e:
                flash(f'下载失败: {str(e)}', 'error')
                return redirect(url_for('excel_converter'))
        
        return redirect(url_for('excel_converter'))


    return app

if __name__ == '__main__':
    app = create_app()
    
    # 创建数据库表（仅在开发环境）
    with app.app_context():
        db.create_all()
    
    app.run(debug=app.config['DEBUG'], host='0.0.0.0', port=5000) 